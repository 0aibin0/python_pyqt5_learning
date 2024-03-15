from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import subprocess
import re
import logging
from openpyxl import load_workbook
import time
from ui_shell_tools import Ui_MainWindow
from PyQt5.QtCore import pyqtSignal
import threading


class controller_main(QMainWindow, Ui_MainWindow):
    return_to_main = pyqtSignal()

    def __init__(self, parent=None):

        super(controller_main, self).__init__(parent)
        self.setupUi(self)
        self.parent = parent
        self.setWindowTitle("shell_tools")
        self.cmdss = None
        self.paths = None
        self.splitsymbol = None
        self.echo = 'echo'
        self.historycmd = '/sys/class/display/dsi/dcs_write'
        self.file_path = ''
        self.isswap = False
        self.astxt = False
        self.wb = None
        self.start_index = '1,1'

        self.comboBox.addItem('echo')
        self.comboBox.addItem('cat')
        self.comboBox.addItem('cd')
        self.comboBox.addItem('ls')
        self.comboBox.addItem('自定义')
        self.comboBox.currentIndexChanged.connect(self.activated)
        self.cmdpath.setText(self.historycmd)
        # xuebin.yin add
        self.lab = QLabel("NIKE", self)
        self.statusbar.addPermanentWidget(self.lab)
        self.timer = QTimer()
        self.timer.timeout.connect(self.showTimeCurrent)
        self.timer.start()

        # self.root_thread = threading.Thread(target=self.root_fun())
        # self.root_thread.start()
        self.pushButton_2.clicked.connect(self.adbdevices_fun)
        self.pushButton_3.clicked.connect(self.adbroot_fun)
        self.pushButton_4.clicked.connect(self.adbreboot_fun)
        self.pushButton_5.clicked.connect(self.ylog_fun)
        self.pushButton_6.clicked.connect(self.lcmapkinstall)
        self.pushButton_7.clicked.connect(self.adbrebootbootloader)
        self.pushButton_8.clicked.connect(self.fastbootreboot)
        self.pushButton.clicked.connect(self.return_to_main.emit)

        self.ylogpath.setPlaceholderText("默认路径为D:\Desktop\ylog")

    def showTimeCurrent(self):
        d = QDateTime.currentDateTime()
        text = d.toString("yyyy-MM-dd HH:mm:ss")
        self.statusbar.showMessage(text, 0)

    def adbdevices_fun(self):
        adbroot = 'adb devices'
        result, out = self.run_command_line(adbroot)
        self.textBrowser.setPlainText(str(result))
        self.textBrowser.setPlainText(str(out))

    def adbroot_fun(self):
        adbdevices = 'adb root'
        result, out = self.run_command_line(adbdevices)
        self.textBrowser.setPlainText(str(out))
        # self.textBrowser.setPlainText("adb root successfully")

    def adbreboot_fun(self):
        adbreboot = 'adb reboot'
        result, out = self.run_command_line(adbreboot)
        # self.textBrowser.setPlainText(str(out))
        self.textBrowser.setPlainText("adb reboot successfully")

    def lcmapkinstall(self):
        lcmapk = 'adb install D:\\00_project\lcm_test_apk\Display-Tester.apk'
        result, out = self.run_command_line(lcmapk)
        self.textBrowser.setPlainText(str(out))
        # self.textBrowser.setPlainText("adb install lcm_test apk successfully")

    def ylog_fun(self):
        ylogpath = self.ylogpath.text()
        if not ylogpath.strip():
            QMessageBox.information(self, "提示", "请填入ylog存储位置", QMessageBox.Ok)
        else:
            adbylog = "adb pull /data/ylog/ap D:\Desktop\ylog\\" + str(ylogpath)
            result, out = self.run_command_line(adbylog)
            self.textBrowser.setPlainText(f"ylog 导出路径为\nD:\Desktop\ylog\\{str(ylogpath)}")

    def adbrebootbootloader(self):
        bootloader = 'adb reboot bootloader'
        result, out = self.run_command_line(bootloader)
        # self.textBrowser.setPlainText(str(out))
        self.textBrowser.setPlainText("enter fastboot mode successfully")

    def fastbootreboot(self):
        bootloader_reboot = 'fastboot reboot'
        result, out = self.run_command_line(bootloader_reboot)
        self.textBrowser.setPlainText(str(out))
        # self.textBrowser.setPlainText("adb install lcm_test apk successfully")

    @pyqtSlot()
    def on_send_clicked(self):
        self.cmdss = self.cmds.text()
        self.paths = self.cmdpath.text()
        self.echo = self.comboBox.currentText()
        if self.echo == 'echo':
            commandline = 'adb shell "' + self.echo + ' ' + self.cmdss + ' > %s' % self.paths + '"'
        else:
            commandline = 'adb shell "' + self.echo + ' %s' % self.paths + '"'
        # self.run_command_line('adb root')
        result, out = self.run_command_line(commandline)

        self.append_text(str(out))

    @pyqtSlot()
    def on_openfile_clicked(self):
        self.file_path = self.dialog()
        self.filepath.setText(self.file_path)

    @pyqtSlot()
    def on_startconvert_clicked(self):
        self.textBrowser.clear()
        try:

            self.wb = load_workbook(filename=self.file_path)
            sheet1 = self.wb.worksheets[0]
            maxrow_s1 = sheet1.max_row
        except Exception as e:
            self.append_text(str(e))
            return
        header = "ocv-capacity-table-0 ="
        result_list = []
        style_result_list = []

        try:
            if not self.startindex.text().strip() == '':
                self.start_index = self.startindex.text()
            start_row = int(self.start_index.split(',')[0])
            start_col = int(self.start_index.split(',')[1])
        except Exception as e:
            self.append_text(str(e))
            return

        try:
            if self.SWAP.isChecked() == True:
                capindex = start_col + 1
                percentindex = start_col
            else:
                capindex = start_col
                percentindex = start_col + 1
            for l in range(start_row, maxrow_s1 + 1):
                capcity = sheet1.cell(row=l, column=capindex).value
                percent = sheet1.cell(row=l, column=percentindex).value
                result = '<' + str(capcity * 1000) + ' ' + '%.0f' % float(percent * 100) + '>'
                result_list.append(result)

        except Exception as e:
            self.append_text(str(e))
            return

        for i in range(len(result_list)):
            if i % 3 == 0 and i > 0:

                style_result_list.append('\n' + ' ' * len(header) + result_list[i])
            else:
                style_result_list.append(result_list[i])

        if self.OUTPUTASFILE.isChecked() == True:
            f = open("output%s.txt" % time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())), "x")
            f.write(header + ','.join(style_result_list) + ';')
            f.close()
        self.append_text(header + ','.join(style_result_list) + ';')

    def run_command_line(self, command_line='NO_COMMAND_GIVEN'):

        P = subprocess.Popen(command_line, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=True)
        out = P.communicate()[0]
        out = out.decode()
        out = re.sub(r'\r+\n', r'\n', out)

        # logging.debug("Command '{0}' output: '{1}'".format(command_line, out))
        result = P.wait()
        if result != 0:
            logging.warning("Command '{0}' failed to run, output:{1}".format(command_line, out))
        return result, out

    def activated(self, index):
        self.echo = self.comboBox.currentText()

    def append_text(self, msg):

        self.textBrowser.insertPlainText(msg)
        self.textBrowser.moveCursor(QTextCursor.End)

    def dialog(self):
        file, check = QFileDialog.getOpenFileName(None, "QFileDialog.getOpenFileName()",
                                                  "", "All Files (*);;Excel Files (*.xlsx)")
        if check:
            return file
        else:
            return ''
